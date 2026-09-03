import csv
import json
from collections import OrderedDict
from collections.abc import Iterable
from io import BytesIO, StringIO
from typing import Any

from openpyxl import load_workbook
from pydantic import ValidationError

from app.core.config import settings
from app.importers._common import import_error, validation_reason
from app.schemas.import_product import NormalizedProduct, NormalizedVariant
from app.schemas.product import ImportError, canonicalize_product_source


def _decode_csv(file_content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8"):
        try:
            return file_content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("CSV file must be encoded as UTF-8 or UTF-8-BOM")


def _parse_json_cell(value: Any, default: Any) -> Any:
    if value is None or value == "":
        return default
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(str(value))
    except json.JSONDecodeError:
        return value


def _parse_tags(value: Any) -> list[str]:
    parsed = _parse_json_cell(value, [])
    if isinstance(parsed, list):
        return [str(item) for item in parsed]
    text = str(value or "")
    delimiter = ";" if ";" in text else "|" if "|" in text else ","
    return [item.strip() for item in text.split(delimiter) if item.strip()]


def _clean_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        str(key).strip().lower(): value.strip() if isinstance(value, str) else value
        for key, value in row.items()
        if key is not None
    }


def _group_tabular_rows(
    rows: Iterable[tuple[int, dict[str, Any]]],
) -> tuple[list[NormalizedProduct], list[ImportError]]:
    grouped: OrderedDict[str, dict[str, Any]] = OrderedDict()
    errors: list[ImportError] = []
    row_count = 0

    for row_number, raw_row in rows:
        row_count += 1
        if row_count > settings.max_import_rows:
            errors.append(
                import_error(
                    row_number,
                    "MAX_IMPORT_ROWS_EXCEEDED",
                    f"Import supports at most {settings.max_import_rows} rows",
                )
            )
            break

        row = _clean_row(raw_row)
        if not any(value not in (None, "") for value in row.values()):
            continue

        external_product_id = str(row.get("external_product_id") or "").strip() or None
        group_key = str(row.get("product_group_key") or "").strip() or None
        if external_product_id is None and group_key is None:
            errors.append(
                import_error(
                    row_number,
                    "PRODUCT_GROUP_KEY_REQUIRED",
                    "File rows require external_product_id or product_group_key",
                )
            )
            continue

        identity = external_product_id or group_key
        assert identity is not None
        try:
            source = canonicalize_product_source(
                str(row.get("source") or "manual")
            )
        except (TypeError, ValueError) as exc:
            errors.append(
                import_error(
                    row_number,
                    "VALIDATION_ERROR",
                    f"source: {exc}",
                )
            )
            continue

        external_shop_id = str(row.get("external_shop_id") or "").strip() or None
        compound_key = f"{source}\x1f{external_shop_id or ''}\x1f{identity}"
        variant_payload = {
            "external_variant_id": row.get("external_variant_id")
            or row.get("variant_external_id"),
            "sku": row.get("sku"),
            "name": row.get("variant_name"),
            "attributes": _parse_json_cell(row.get("attributes"), {}),
            "price": row.get("price") or None,
            "original_price": row.get("original_price") or None,
            "stock_quantity": row.get("stock_quantity") or None,
            "stock_status": row.get("stock_status") or "unknown",
            "image_url": row.get("variant_image_url"),
            "metadata": _parse_json_cell(row.get("variant_metadata"), {}),
            "source_updated_at": row.get("variant_source_updated_at"),
        }
        try:
            variant = NormalizedVariant.model_validate(variant_payload)
        except ValidationError as exc:
            errors.append(
                import_error(
                    row_number,
                    "VALIDATION_ERROR",
                    validation_reason(exc),
                )
            )
            continue

        if compound_key not in grouped:
            grouped[compound_key] = {
                "row_number": row_number,
                "source": source,
                "external_product_id": external_product_id or group_key,
                "external_shop_id": external_shop_id,
                "name": row.get("name"),
                "description": row.get("description"),
                "category": row.get("category"),
                "tags": _parse_tags(row.get("tags")),
                "currency": row.get("currency") or "VND",
                "status": row.get("status") or "active",
                "image_url": row.get("image_url"),
                "product_url": row.get("product_url"),
                "affiliate_url": row.get("affiliate_url"),
                "metadata": _parse_json_cell(row.get("metadata"), {}),
                "source_updated_at": row.get("source_updated_at"),
                "variants": [],
            }
        grouped[compound_key]["variants"].append(variant)

    products: list[NormalizedProduct] = []
    for payload in grouped.values():
        row_number = payload.pop("row_number")
        try:
            products.append(NormalizedProduct.model_validate(payload))
        except ValidationError as exc:
            errors.append(
                import_error(
                    row_number,
                    "VALIDATION_ERROR",
                    validation_reason(exc),
                )
            )
    return products, errors


def parse_csv_import(
    file_content: bytes,
) -> tuple[list[NormalizedProduct], list[ImportError]]:
    try:
        decoded = _decode_csv(file_content)
    except ValueError as exc:
        return [], [import_error(1, "INVALID_ENCODING", str(exc))]
    if not decoded.strip():
        return [], [import_error(1, "EMPTY_FILE", "CSV file is empty")]

    reader = csv.DictReader(StringIO(decoded))
    if reader.fieldnames is None:
        return [], [
            import_error(1, "MISSING_HEADER", "CSV header row is required")
        ]
    return _group_tabular_rows((reader.line_num, row) for row in reader)


def parse_xlsx_import(
    file_content: bytes,
) -> tuple[list[NormalizedProduct], list[ImportError]]:
    try:
        workbook = load_workbook(
            BytesIO(file_content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:  # noqa: BLE001
        return [], [
            import_error(1, "INVALID_XLSX", f"Unable to read XLSX: {exc}")
        ]

    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    headers = next(iterator, None)
    if headers is None:
        workbook.close()
        return [], [import_error(1, "EMPTY_FILE", "XLSX file is empty")]

    normalized_headers = [str(value or "").strip().lower() for value in headers]
    rows = (
        (row_number, dict(zip(normalized_headers, values, strict=False)))
        for row_number, values in enumerate(iterator, start=2)
    )
    try:
        return _group_tabular_rows(rows)
    finally:
        workbook.close()
